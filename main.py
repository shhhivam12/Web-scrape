from flask import Flask, request, render_template, make_response
import bs4 as bs
import requests
import pandas as pd
import openpyxl
import io
import base64
import pickle

app = Flask(__name__)

@app.route('/')
def home():
    return render_template('home.html')

@app.route('/scrape', methods=['POST'])
def scrape():
    if request.method == 'POST':
        url = request.form['url']
        page = requests.get(url)
        page_data = page.text
        soup = bs.BeautifulSoup(page_data, features="html.parser")
        all_tables = soup.findAll('table')

        tables_data = []
        limited_tables_html = []
        for table in all_tables:
            table_html = str(table)
            tables_data.append(table_html)
            df = pd.read_html(io.StringIO(table_html))[0]
            limited_df = df.head(5)
            limited_table_html = limited_df.to_html(index=False)
            limited_tables_html.append(limited_table_html)

        encoded_tables_data = base64.b64encode(pickle.dumps(tables_data)).decode('utf-8')

        return render_template('home.html', scraped_tables=limited_tables_html, tables_data=encoded_tables_data)

@app.route('/download_excel', methods=['POST'])
def download_excel():
    table_index = int(request.form['table_index'])
    tables_data = pickle.loads(base64.b64decode(request.form['tables_data']))

    table_html = tables_data[table_index]
    soup = bs.BeautifulSoup(table_html, features="html.parser")
    table = soup.find('table')

    wb = openpyxl.Workbook()
    ws = wb.active

    for row_idx, row in enumerate(table.find_all('tr'), 1):
        for col_idx, cell in enumerate(row.find_all(['td', 'th']), 1):
            if ws.cell(row=row_idx, column=col_idx).value is None:  # Check if the cell is already part of a merged cell
                cell_value = cell.get_text()
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
                if cell.has_attr('colspan'):
                    col_span = int(cell['colspan'])
                    ws.merge_cells(start_row=row_idx, start_column=col_idx, end_row=row_idx, end_column=col_idx + col_span - 1)
                if cell.has_attr('rowspan'):
                    row_span = int(cell['rowspan'])
                    ws.merge_cells(start_row=row_idx, start_column=col_idx, end_row=row_idx + row_span - 1, end_column=col_idx)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=table.xlsx"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response

if __name__ == '__main__':
    app.run(debug=True)
